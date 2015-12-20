# -*- coding: utf-8 -*-
"""
/***************************************************************************
 UniumPlugin
                                 A QGIS plugin
 The plugin for Unium
                              -------------------
        begin                : 2015-11-30
        git sha              : $Format:%H$
        copyright            : (C) 2015 by Ivan Medvedev/MapCrap
        email                : medvedev.ivan@mail.ru
 ***************************************************************************/

/***************************************************************************
 *                                                                         *
 *   This program is free software; you can redistribute it and/or modify  *
 *   it under the terms of the GNU General Public License as published by  *
 *   the Free Software Foundation; either version 2 of the License, or     *
 *   (at your option) any later version.                                   *
 *                                                                         *
 ***************************************************************************/
"""
from PyQt4.QtCore import QSettings, QTranslator, qVersion, QCoreApplication, Qt, QVariant
from PyQt4.QtGui import QAction, QIcon, QFileDialog, QTableWidgetItem, qApp, QImage, QPixmap
from PyQt4.QtCore import pyqtSlot,SIGNAL,SLOT
from qgis.core import *
from qgis.gui import *
from lxml import etree
from openpyxl import Workbook, load_workbook
# Initialize Qt resources from file resources.py
import resources, os, sqlite3, shutil, datetime, json, math, sys

# Import the code for the DockWidget
from mcqp_unium_dockwidget import UniumPluginDockWidget
import os.path
from __builtin__ import isinstance


class UniumPlugin:
    """QGIS Plugin Implementation."""

    def __init__(self, iface):
        """Constructor.

        :param iface: An interface instance that will be passed to this class
            which provides the hook by which you can manipulate the QGIS
            application at run time.
        :type iface: QgsInterface
        """
        # Save reference to the QGIS interface
        self.iface = iface

        # initialize plugin directory
        self.plugin_dir = os.path.dirname(__file__)

        # initialize locale
        locale = QSettings().value('locale/userLocale')[0:2]
        locale_path = os.path.join(
            self.plugin_dir,
            'i18n',
            'UniumPlugin_{}.qm'.format(locale))

        if os.path.exists(locale_path):
            self.translator = QTranslator()
            self.translator.load(locale_path)

            if qVersion() > '4.3.3':
                QCoreApplication.installTranslator(self.translator)

        # Declare instance attributes
        self.actions = []
        self.menu = self.tr(u'&UniumPlugin')
        # TODO: We are going to let the user set this up in a future iteration
        self.toolbar = self.iface.pluginToolBar() #self.iface.addToolBar(u'UniumPlugin')
        self.toolbar.setObjectName(u'UniumPlugin')

        #print "** INITIALIZING UniumPlugin"

        self.pluginIsActive = False
        self.dockwidget = None
        
        # initialize dict for saving data about importing sml
        self.sml_data = {}
        
        # initialize layers data
        self.categories = {}
        self.layers = {}
        self.src_info = {}
        self.selected_id = u''
        self.mercator = QgsCoordinateReferenceSystem()
        mercatorWKT = u'PROJCS["WGS 84 / Pseudo-Mercator",GEOGCS["WGS 84",DATUM["WGS_1984",SPHEROID["WGS 84",6378137,298.257223563,AUTHORITY["EPSG","7030"]],AUTHORITY["EPSG","6326"]],PRIMEM["Greenwich",0,AUTHORITY["EPSG","8901"]],UNIT["degree",0.0174532925199433,AUTHORITY["EPSG","9122"]],AUTHORITY["EPSG","4326"]],PROJECTION["Mercator_1SP"],PARAMETER["central_meridian",0],PARAMETER["scale_factor",1],PARAMETER["false_easting",0],PARAMETER["false_northing",0],UNIT["metre",1,AUTHORITY["EPSG","9001"]],AXIS["X",EAST],AXIS["Y",NORTH],EXTENSION["PROJ4","+proj=merc +a=6378137 +b=6378137 +lat_ts=0.0 +lon_0=0.0 +x_0=0.0 +y_0=0 +k=1.0 +units=m +nadgrids=@null +wktext  +no_defs"],AUTHORITY["EPSG","3857"]]'
        mercatorProj4 = u'+proj=merc +a=6378137 +b=6378137 +lat_ts=0.0 +lon_0=0.0 +x_0=0.0 +y_0=0 +k=1.0 +units=m +nadgrids=@null +wktext  +no_defs'
        #self.mercator.createFromProj4()
        self.mercator.createFromWkt(mercatorWKT)

        # default configuration
        self.default_config = """{"files_folder": "",
                            "write_block": 50,
                            "signs": {"school_2zone": {"alias": u"Школа (2 зона)", "png_src": ["blue.png"], "filename": "circle_house_blue"},
                                      "school_3zone": {"alias": u"Школа (3 зона)", "png_src": ["green.png"], "filename": "circle_house_green"},
                                      "school_1zone": {"alias": u"Школа (1 зона)", "png_src": ["red.png"], "filename": "circle_house_red"},
                                      "school_repeat": {"alias": u"Школа (повтор)", "png_src": ["dopprohod.png","doprohod.png","yellow.png"], "filename": "circle_house_yellow"},
                                      "school_center": {"alias": u"Образовательный центр", "png_src": ["stationnew.png"], "filename": "circle_house_black"},
                                      "school_bad": {"alias": u"Не подходит", "png_src": ["neproshel.png"], "filename": "circle_cross2"},
                                      "school_stop": {"alias": u"Не пустили", "png_src": ["nepustili.png"], "filename": "circle_minus2"},
                                      "school_adv": {"alias": u"Объявления", "png_src": ["obyavlenie.png"], "filename": "triangle_warning2"},
                                      "school_passed": {"alias": u"Пройдена", "png_src": ["proshel.png"], "filename": "circle_checktick2"},
                                      "school_atwork": {"alias": u"Выдана агенту", "png_src": ["vydanaagentu.png"], "filename": "circle_helmet"},
                                      "school_freepass": {"alias": u"Пройдена без договорённостей", "png_src": ["climbing.png"], "filename": "circle_climbing"},
                                      "other_tools": {"alias": u"[Ремонт]", "png_src": ["carrepair.png"], "filename": "square_tools"},
                                      "other_house": {"alias": u"[Дом]", "png_src": ["home.png"], "filename": "square_house_black"},
                                      "other_stop": {"alias": u"[Стоп]", "png_src": ["stop.png"], "filename": "square_stop"},
                                      "other_stoplight": {"alias": u"[Светофор]", "png_src": ["stoplight.png"], "filename": "square_trafficlights"},
                                      "other_unknown" : {"alias" : "[без знака]", "filename" : "circle_unknown_grey"},
                                      "promo_red": {"alias": u"Промо (срочно)", "png_src": ["promogive.png"], "filename": "circle_pushpin_red"},
                                      "promo_yellow": {"alias": u"Промо (обычный)", "png_src": ["promoter.png"], "filename": "circle_pushpin_yellow"},
                                      "promo_green": {"alias": u"Промо (хороший)", "png_src": ["promotergood.png"], "filename": "circle_pushpin_green"}
                            }
                            """


    # noinspection PyMethodMayBeStatic
    def tr(self, message):
        """Get the translation for a string using Qt translation API.

        We implement this ourselves since we do not inherit QObject.

        :param message: String for translation.
        :type message: str, QString

        :returns: Translated version of message.
        :rtype: QString
        """
        # noinspection PyTypeChecker,PyArgumentList,PyCallByClass
        return QCoreApplication.translate('UniumPlugin', message)


    def add_action(
        self,
        icon_path,
        text,
        callback,
        enabled_flag=True,
        add_to_menu=True,
        add_to_toolbar=True,
        status_tip=None,
        whats_this=None,
        parent=None):
        """Add a toolbar icon to the toolbar.

        :param icon_path: Path to the icon for this action. Can be a resource
            path (e.g. ':/plugins/foo/bar.png') or a normal file system path.
        :type icon_path: str

        :param text: Text that should be shown in menu items for this action.
        :type text: str

        :param callback: Function to be called when the action is triggered.
        :type callback: function

        :param enabled_flag: A flag indicating if the action should be enabled
            by default. Defaults to True.
        :type enabled_flag: bool

        :param add_to_menu: Flag indicating whether the action should also
            be added to the menu. Defaults to True.
        :type add_to_menu: bool

        :param add_to_toolbar: Flag indicating whether the action should also
            be added to the toolbar. Defaults to True.
        :type add_to_toolbar: bool

        :param status_tip: Optional text to show in a popup when mouse pointer
            hovers over the action.
        :type status_tip: str

        :param parent: Parent widget for the new action. Defaults None.
        :type parent: QWidget

        :param whats_this: Optional text to show in the status bar when the
            mouse pointer hovers over the action.

        :returns: The action that was created. Note that the action is also
            added to self.actions list.
        :rtype: QAction
        """

        icon = QIcon(icon_path)
        action = QAction(icon, text, parent)
        action.triggered.connect(callback)
        action.setEnabled(enabled_flag)

        if status_tip is not None:
            action.setStatusTip(status_tip)

        if whats_this is not None:
            action.setWhatsThis(whats_this)

        if add_to_toolbar:
            self.toolbar.addAction(action)

        if add_to_menu:
            self.iface.addPluginToMenu(
                self.menu,
                action)

        self.actions.append(action)

        return action


    def initGui(self):
        """Create the menu entries and toolbar icons inside the QGIS GUI."""

        icon_path = ':/plugins/UniumPlugin/icon.png'
        self.add_action(
            icon_path,
            text=self.tr(u'Unium'),
            callback=self.run,
            parent=self.iface.mainWindow())

    #--------------------------------------------------------------------------

    def onClosePlugin(self):
        """Cleanup necessary items here when plugin dockwidget is closed"""

        #print "** CLOSING UniumPlugin"

        self.set_project_settings()

        # disconnects
        self.dockwidget.closingPlugin.disconnect(self.onClosePlugin)

        # remove this statement if dockwidget is to remain
        # for reuse if plugin is reopened
        # Commented next statement since it causes QGIS crashe
        # when closing the docked window:
        # self.dockwidget = None

        self.pluginIsActive = False


    def unload(self):
        """Removes the plugin menu item and icon from QGIS GUI."""

        #print "** UNLOAD UniumPlugin"

        for action in self.actions:
            self.iface.removePluginMenu(
                self.tr(u'&UniumPlugin'),
                action)
            self.iface.removeToolBarIcon(action)
        # remove the toolbar
        del self.toolbar

    #--------------------------------------------------------------------------

    def run(self):
        """Run method that loads and starts the plugin"""
        
        if not self.pluginIsActive:
            self.pluginIsActive = True

            #print "** STARTING UniumPlugin"

            # dockwidget may not exist if:
            #    first run of plugin
            #    removed on close (see self.onClosePlugin method)
            if self.dockwidget == None:
                # Create the dockwidget (after translation) and keep reference
                self.dockwidget = UniumPluginDockWidget()
                self.dockwidget.catbrwsButton.clicked.connect(self.select_cat_file)
                self.dockwidget.marksbrwsButton.clicked.connect(self.select_marks_file)
                self.dockwidget.dbbrwsButton.clicked.connect(self.select_db_file)
                self.dockwidget.loadSASButton.clicked.connect(self.loadSASButton_clicked)
                self.dockwidget.layersBox.currentIndexChanged.connect(self.selected_layer_changed)
                self.dockwidget.ffbrwsButton.clicked.connect(self.select_filefolder_clicked)
                self.dockwidget.wrblkBox.valueChanged.connect(self.wrblkBox_value_changed)
                self.dockwidget.savesetButton.clicked.connect(self.savesetButton_clicked)
                self.dockwidget.loadsetButton.clicked.connect(self.loadsetButton_clicked)
                self.dockwidget.applyFilterButton.clicked.connect(self.applyFilterButton_clicked)
                self.dockwidget.resetFilterButton.clicked.connect(self.resetFilterButton_clicked)
                self.dockwidget.brwsxlsoutButton.clicked.connect(self.select_out_excel_file)
                self.dockwidget.xlsoutButton.clicked.connect(self.export_to_xls)
                self.dockwidget.brwsxlsinButton.clicked.connect(self.select_in_excel_file)
                self.dockwidget.xlsinButton.clicked.connect(self.import_from_xls)
                #self.dockwidget.layersBox.connect(self.dockwidget.layersBox,SIGNAL("currentIndexChanged(int)"),self.dockwidget,SLOT("self.selected_layer_changed(int)"))
                
            self.dockwidget.layersBox.currentIndex = 1
            
            # connect to provide cleanup on closing of dockwidget
            self.dockwidget.closingPlugin.connect(self.onClosePlugin)

            self.getSettings()
            self.updateSettingsUI()
            self.get_project_settings()
            self.update_layers_list()
            for lyr in self.iface.legendInterface().layers():
                if isinstance(lyr, QgsVectorLayer):
                    self.load_subsets(lyr)

            # show the dockwidget
            # TODO: fix to allow choice of dock location
            self.iface.addDockWidget(Qt.BottomDockWidgetArea, self.dockwidget)
            self.dockwidget.show()

    #--------------------------------------------------------------------------

    def getSettings(self):
        config_file = os.path.join(os.path.dirname(__file__),'mcqp_unium_config.json')
        msg = u"Конфигурационный файл отсутствует. Будет загружена конфигурация по-умолчанию"
        if os.path.exists(config_file):
            try:
                conf = open(config_file,'r')
                self.config = json.load(conf)
                conf.close()
                i_msg = u"Конфигурация загружена"
                self.iface.messageBar().pushMessage("Info", i_msg, level=QgsMessageBar.INFO, duration=7)
                QgsMessageLog.logMessage(i_msg, level=QgsMessageLog.INFO)
                return
            except Exception,err:
                msg = u"Ошибка при загрузке конфигурационного файла: %s. Будет загружена конфигурация по-умолчанию" % err
                self.config = self.default_config
        self.iface.messageBar().pushMessage("Warning", msg, level=QgsMessageBar.WARNING, duration=7)
        QgsMessageLog.logMessage(msg, level=QgsMessageLog.WARNING)
        return

    def setSettings(self):
        config_file = os.path.join(os.path.dirname(__file__),'mcqp_unium_config.json')
        if os.path.exists(config_file):
            try:
                conf = open(config_file,'w')
                json.dump(self.config,conf)
                conf.close()
                i_msg = u"Конфигурация сохранена"
                self.iface.messageBar().pushMessage("Info", i_msg, level=QgsMessageBar.INFO, duration=7)
                QgsMessageLog.logMessage(i_msg, level=QgsMessageLog.INFO)
                return
            except Exception,err:
                msg = u"Ошибка при записи конфигурационного файла: %s." % err
                self.iface.messageBar().pushMessage("Warning", msg, level=QgsMessageBar.WARNING, duration=7)
                QgsMessageLog.logMessage(msg, level=QgsMessageLog.WARNING)

    def get_project_settings(self):
        # Читаем список категорий в формате {<id>:"<категория>"}
        self.categories = json.loads(QgsProject.instance().readEntry("UniumPlugin", "categories", "{}")[0])
        self.src_info["database"] = QgsProject.instance().readEntry("UniumPlugin", "database", "")[0]
        self.src_info["datatable"] = QgsProject.instance().readEntry("UniumPlugin", "datatable", "")[0]

    def set_project_settings(self):
        if QgsProject:
            QgsProject.instance().writeEntry("UniumPlugin", "categories", json.dumps(self.categories))
            QgsProject.instance().writeEntry("UniumPlugin", "database", self.src_info.get("database",""))
            QgsProject.instance().writeEntry("UniumPlugin", "datatable", self.src_info.get("datatable",""))

    def updateSettingsUI(self):
        self.dockwidget.filefolderEdit.setText(self.config['files_folder'])
        self.dockwidget.wrblkBox.setValue(self.config['write_block'])

    #--------------------------------------------------------------------------

    def get_sign_by_src(self,png_src):
        for sign_key in self.config.get("signs",{}).keys():
            if png_src in self.config["signs"][sign_key].get("png_src",[]):
                return sign_key
        return 'other_unknown'

    def get_sign_by_alias(self,alias):
        for sign_key in self.config.get("signs",{}).keys():
            if alias == self.config["signs"][sign_key].get("alias",""):
                return sign_key
        return 'other_unknown'

    def get_sign_image(self,sign):
        img_path = os.path.join(self.plugin_dir,'images','png','%s.png' % self.config["signs"].get(sign,{}).get("filename",'image'))
        if os.path.exists(img_path):
            image = QImage()
            image.load(img_path)
            pm = QPixmap.fromImage(image)
            return pm
        else:
            return False


    #--------------------------------------------------------------------------

    def export_to_xls(self):
        if self.dockwidget.xlsoutEdit.text():
            try:
                xls_filename = self.dockwidget.xlsoutEdit.text()
                if os.path.exists(xls_filename):
                    try:
                        os.remove(xls_filename)
                    except Exception, err:
                        msg = u"Ошибка при удалении файла %s : %s" % (unicode(xls_filename),err)
                        self.iface.messageBar().pushMessage("Error", msg, level=QgsMessageBar.CRITICAL, duration=7)
                        QgsMessageLog.logMessage(msg, level=QgsMessageLog.CRITICAL)
                        return
                if self.selected_id in self.layers.keys():
                    wb = Workbook()
                    ws = wb.active
                    self.dockwidget.layersBox.enabled = False
                    ws.append([u"Идентификатор",u"Наименование",u"Описание",u"Условный знак",u"Широта",u"Долгота",u"Номер категории",u"Категория"])
                    lyrs = self.iface.legendInterface().layers()
                    for lyr in lyrs:
                        all_case = (self.selected_id == '0' and self.iface.legendInterface().isLayerVisible(lyr))
                        if isinstance(lyr, QgsVectorLayer) and (lyr.id() == self.selected_id or all_case):
                            features = lyr.getFeatures()
                            for feat in features:
                                geom = feat.geometry()
                                ws.append([feat["id"],feat["name"],feat["descr"],self.config.get("signs",{}).get(feat["sign"],{}).get("alias",""),geom.asPoint().y(), geom.asPoint().x(),feat["cat_id"],self.categories.get(unicode(feat["cat_id"]),u'')])
                            if not all_case:
                                break
                    wb.save(xls_filename)
                    msg = u"Выгрузка завершена"
                    if sys.platform.startswith('win'):
                        os.system("start "+xls_filename)
                self.iface.messageBar().pushMessage(u"Выгрузка в Excel", msg, level=QgsMessageBar.INFO, duration=7)
                QgsMessageLog.logMessage(msg, level=QgsMessageLog.INFO)
            except Exception, err:
                msg = u"Ошибка при выгрузке в Excel: %s" % err
                self.iface.messageBar().pushMessage("Error", msg, level=QgsMessageBar.CRITICAL, duration=7)
                QgsMessageLog.logMessage(msg, level=QgsMessageLog.CRITICAL)

    def import_from_xls(self):
        if self.dockwidget.xlsinEdit.text():
            try:
                # Open xls
                xls_filename = self.dockwidget.xlsinEdit.text()
                wb = load_workbook(filename = xls_filename)
                ws = wb.active

                # Prepare work layer
                uri = QgsDataSourceURI()
                uri.setDatabase(self.src_info['database'])
                uri.setDataSource('', self.src_info['datatable'], 'shape')

                lyr = QgsVectorLayer(uri.uri(),self.src_info['datatable'],'spatialite')
                pr = lyr.dataProvider()
                lyr.startEditing()

                # Checking for new categories
                new_cats = {}
                max_catid = self.get_max_catid()
                c_catid = max_catid+1

                for row in ws.rows[1:]:
                    if not row[6].value:
                        if row[7].value in self.categories.values():
                            new_cats[row[7].value] = self.get_catid_by_path(row[7].value)
                        else:
                            new_cats[row[7].value] = c_catid
                            c_catid+=1

                # Iterate rows
                check_nulls = lambda val: val if val else ''
                for row in ws.rows[1:]:
                    drow = {}
                    for i,cell in enumerate(row):
                        drow[i] = cell.value
                    # If id field not empty - try update it
                    if drow[0]:
                        lyr.beginEditCommand("Feature update")
                        try:
                            f_cat_id = drow[6]
                            if not drow[6]:
                                f_cat_id = new_cats[drow[7]]
                            attrs = { 1 : check_nulls(drow[1]), 2 : check_nulls(drow[2]), 3: self.get_sign_by_alias(drow[3]), 4: f_cat_id}
                            geom = QgsGeometry.fromPoint(QgsPoint(drow[5],drow[4]))
                            pr.changeAttributeValues({ drow[0] : attrs })
                            pr.changeGeometryValues({ drow[0] : geom })
                            lyr.endEditCommand()
                        except Exception,err:
                            lyr.destroyEditCommand()
                            QgsMessageLog.logMessage(u'Ошибка при изменении метки id=%s: %s' % (drow[0],err), level=QgsMessageLog.CRITICAL)

                    # If id field is empty - try insert new mark
                    else:
                        lyr.beginEditCommand("Feature insert")
                        try:
                            feature = QgsFeature()
                            feature.setGeometry(QgsGeometry.fromPoint(QgsPoint(drow[5],drow[4])))
                            fields = pr.fields()
                            feature.setFields(fields)
                            f_cat_id = drow[6]
                            if not f_cat_id:
                                f_cat_id = new_cats[drow[7]]
                            feature.setAttribute('name',check_nulls(drow[1]))
                            feature.setAttribute('descr',check_nulls(drow[2]))
                            feature.setAttribute('sign',self.get_sign_by_alias(drow[3]))
                            feature.setAttribute('cat_id',f_cat_id)
                            pr.addFeatures([feature])
                            lyr.endEditCommand()
                        except Exception,err:
                            lyr.destroyEditCommand()
                            QgsMessageLog.logMessage(u'Ошибка при добавлении метки: %s' % err, level=QgsMessageLog.CRITICAL)
                lyr.commitChanges()
                lyr = None
                wb.save(xls_filename)

                root = QgsProject.instance().layerTreeRoot()

                # Create and register new sublayers
                for cat in new_cats.keys():
                    chain = cat.split(chr(92))
                    self.categories[new_cats[cat]] = cat
                    # Create sublayers for category
                    c_root = UniumPlugin.create_sublayers(root, chain)
                    # Create category vector layer
                    cat_lyr = self.create_catlyr(uri.uri(),chain[len(chain)-1],new_cats[cat])
                    self.layers[cat_lyr.id()] = {'name': cat_lyr.name(),
                                        'subset': cat_lyr.subsetString(),
                                        'path': chr(92).join(chain[:len(chain)-1]),
                                        'full_name':chr(92).join(chain)}
                    QgsMapLayerRegistry.instance().addMapLayer(cat_lyr,False)
                    c_root.addLayer(cat_lyr)
                    self.iface.legendInterface().setLayerVisible(cat_lyr, False)

                self.set_project_settings()
                self.iface.mapCanvas().mapRenderer().setDestinationCrs(QgsCoordinateReferenceSystem(3857, QgsCoordinateReferenceSystem.EpsgCrsId))
                self.iface.mapCanvas().setMapUnits(0)
                self.set_style_to_lyrs()
                self.update_TView()
                self.update_layers_list()
                msg = u"Загрузка завершена"
                self.iface.messageBar().pushMessage(u"Загрузка в Excel", msg, level=QgsMessageBar.INFO, duration=7)
                QgsMessageLog.logMessage(msg, level=QgsMessageLog.INFO)

            except Exception, err:
                msg = u"Ошибка при загрузке в Excel: %s" % err
                self.iface.messageBar().pushMessage("Error", msg, level=QgsMessageBar.CRITICAL, duration=7)
                QgsMessageLog.logMessage(msg, level=QgsMessageLog.CRITICAL)

    #--------------------------------------------------------------------------

    @staticmethod
    def rec_get_layer_path(root,layer_id,path):
        check = False
        for node in root.children():
            if isinstance(node,QgsLayerTreeGroup):
                (check,path) = UniumPlugin.rec_get_layer_path(node,layer_id,path)
                if check:
                    path.append(root.name())
                    return (check,path)
            elif isinstance(node,QgsLayerTreeLayer):
                if node.layerId() == layer_id:
                    path.append(root.name())
                    check = True
        return (check,path)
        
    @staticmethod
    def get_layer_path(root,layer_id):
        path = []
        for node in root.children():
            if isinstance(node,QgsLayerTreeGroup):
                check = False
                (check,path) = UniumPlugin.rec_get_layer_path(node,layer_id,path)
                if check:
                    path.append(root.name())
                    break
            elif isinstance(node,QgsLayerTreeLayer):
                if node.layerId() == layer_id:
                    path.append(root.name())
        path.reverse()
        return chr(92).join([p for p in path if len(p) > 0])

    def get_layers(self):
        self.layers = {"0": {"name": u"Все", "path":"", "full_name": u"Все"}}
        for lyr in self.iface.legendInterface().layers():
            if isinstance(lyr, QgsVectorLayer):
                path = UniumPlugin.get_layer_path(QgsProject.instance().layerTreeRoot(),lyr.id())
                self.layers[lyr.id()] = {'name': lyr.name(),
                                    'subset': lyr.subsetString(),
                                    'path': path,
                                    'full_name':chr(92).join([path,lyr.name()])}
    
    @staticmethod
    def __find_same_layers(root,lyr_ids):
        group_names = [l.name for l in root.children() if isinstance(l, QgsLayerTreeGroup)]
        for tl in root.children():
            if isinstance(tl, QgsLayerTreeGroup):
                lyr_ids = UniumPlugin.__find_same_layers(tl, lyr_ids)
            elif isinstance(tl, QgsLayerTreeLayer):
                if tl.layerName() in group_names:
                    lyr_ids.append(tl.layerId())
        return lyr_ids
                
    def remove_vlayers_as_group(self):
        root = QgsProject.instance().layerTreeRoot()
        ids = []
        ids = UniumPlugin.__find_same_layers(root, ids)
        if len(ids) > 0:
            lyrs = self.iface.legendInterface().layers()
            for lyr in lyrs:
                if isinstance(lyr, QgsVectorLayer) and lyr.id() in ids:
                    l_feats = lyr.getFeatures()
                    if len(l_feats) <= 0:
                        QgsMapLayerRegistry.instance().removeMapLayer(lyr)
                        self.layers.pop(lyr.id())
                                    
    def set_style_to_lyrs(self):
        lyrs = self.iface.legendInterface().layers()
        for lyr in lyrs:
            if isinstance(lyr, QgsVectorLayer) and lyr.id() in self.layers.keys():
                lyr.setDisplayField('name')
                lyr_qml = os.path.join(self.plugin_dir,'lyr.qml')
                if os.path.exists(lyr_qml):
                    lyr.loadNamedStyle(lyr_qml)
                lyr_lbl = os.path.join(self.plugin_dir,'lyr_label.conf')
                if os.path.exists(lyr_lbl):
                    with open(lyr_lbl,'r') as lyr_lbl_f:
                        lyr_lbl_conf = json.load(lyr_lbl_f)
                        for lkey in lyr_lbl_conf.keys():
                            lyr.setCustomProperty(lkey,lyr_lbl_conf[lkey])

    def update_layers_list(self):
        list_items = []
        self.dockwidget.layersBox.clear()
        self.get_layers()
        for lid in self.layers.keys():
            list_items.append(self.layers[lid].get('full_name',''))
        list_items.sort()
        self.dockwidget.layersBox.addItems(list_items)

    def update_TView(self):
        if self.selected_id in self.layers.keys():
            self.dockwidget.layersBox.enabled = False
            attrs_names = [u"Идентификатор",u"Наименование",u"Описание",u"Условный знак"]
            attrs_values = []
            lif = self.iface.legendInterface()
            lyrs = lif.layers()
            for lyr in lyrs:
                all_case = (self.selected_id == '0' and lif.isLayerVisible(lyr))
                if isinstance(lyr, QgsVectorLayer) and (lyr.id() == self.selected_id or all_case):
                    #attrs_names = [a.name() for a in lyr.fields()]
                    features = lyr.getFeatures()
                    attrs_values += [[feat[i] for i in xrange(len(attrs_names))] for feat in features]
                    if not all_case:
                        break
            self.dockwidget.tableView.setRowCount(len(attrs_values))
            self.dockwidget.tableView.setColumnCount(len(attrs_names))
            self.dockwidget.tableView.setHorizontalHeaderLabels(attrs_names)
            for row in xrange(len(attrs_values)):
                for col in xrange(len(attrs_names)):
                    if col < len(attrs_names) - 1:
                        item = QTableWidgetItem(u'%s' % attrs_values[row][col])
                    else:
                        pm = self.get_sign_image(attrs_values[row][col])
                        if pm:
                            item = QTableWidgetItem(self.config["signs"][attrs_values[row][col]].get("alias",""))
                            item.setData(Qt.DecorationRole, pm.scaled(20, 20))
                        else:
                            item = QTableWidgetItem(u'%s' % attrs_values[row][col])
                    self.dockwidget.tableView.setItem(row,col,item)
            self.dockwidget.layersBox.enabled = True

    # create lyr for category
    def create_catlyr(self,uri,chain,cat_id):
        cat_lyr = QgsVectorLayer(uri, chain, 'spatialite')
        cat_lyr.setCustomProperty("cat_filter", cat_id)
        self.reset_subsets(cat_lyr)
        return cat_lyr

    def get_catid_by_path(self,path):
        for catid in self.categories.keys():
            if self.categories[catid] == path:
                return int(catid)
        return -1

    @staticmethod
    def create_sublayers(root, chain):
        c_root = root
        for i in xrange(len(chain)-1):
            c_node = c_root.findGroup(chain[i])
            if not c_node:
                c_node = c_root.addGroup(chain[i])
            c_root = c_node
        return c_root

    def set_subsets(self,lyr,name = '',descr = ''):
        subset_str = ''
        cat_filter = lyr.customProperty("cat_filter", -1)
        if cat_filter <> -1:
            subset_str = u'("cat_id" = %s)' % cat_filter
        if name <> '':
            subset_str += u' & ("name" like \'%{0}%\')'.format(name)
            lyr.setCustomProperty("name_filter", name)
            self.layers[lyr.id()]["name_filter"] = name
        if descr <> '':
            subset_str += u' & ("descr" like \'%{0}%\')'.format(descr)
            lyr.setCustomProperty("descr_filter", descr)
            self.layers[lyr.id()]["descr_filter"] = descr
        lyr.setSubsetString(subset_str)

    def reset_subsets(self,lyr):
        subset_str = ''
        cat_filter = lyr.customProperty("cat_filter", -1)
        if cat_filter <> -1:
            subset_str = u'("cat_id" = %s)' % cat_filter
            lyr.setCustomProperty("name_filter", '')
            lyr.setCustomProperty("descr_filter", '')
            if self.layers.has_key(lyr.id()):
                self.layers[lyr.id()]["name_filter"] = ''
                self.layers[lyr.id()]["descr_filter"] = ''
        lyr.setSubsetString(subset_str)

    def load_subsets(self,lyr):
        name_filter = lyr.customProperty("name_filter", '')
        descr_filter = lyr.customProperty("descr_filter", '')
        self.set_subsets(lyr,name_filter,descr_filter)
    
    @staticmethod
    def create_db(db_file):
        con = sqlite3.connect(db_file)
        cur = con.cursor()
        con.enable_load_extension(True)
        con.execute("SELECT load_extension('mod_spatialite.dll')")
        sql = "SELECT InitSpatialMetadata(1,'WGS84')"
        cur.execute(sql)
        con.close()
    
    @staticmethod
    def create_table(db_file,table_name):
        con = sqlite3.connect(db_file)
        cur = con.cursor()
        con.enable_load_extension(True)
        con.execute("SELECT load_extension('mod_spatialite.dll')")
        count = cur.execute("select count(name) from sqlite_master where type='table' and name = ?",(table_name,))
        if count.fetchall()[0][0] > 0:
            table_name += '_%s' % datetime.datetime.now().strftime('%Y%m%d_%H%M')
        cur.execute('CREATE TABLE %s (id INTEGER NOT NULL PRIMARY KEY,name TEXT,descr TEXT,sign TEXT,cat_id INTEGER)'% table_name)
        con.commit()
        cur.execute("SELECT AddGeometryColumn(?,'shape', 4326, 'POINT', 'XY')",(table_name,))
        con.commit()
        con.close()
        return table_name

    def get_max_catid(self):
        con = sqlite3.connect(self.src_info["database"])
        cur = con.cursor()
        res = cur.execute('select max(cat_id) from %s' % self.src_info["datatable"])
        max_id = res.fetchall()[0][0]
        con.close()
        return max_id
    
    def ParseSML(self):
        """Parse SAS Planet data to splite db and make layers tree"""
        try:
            self.dockwidget.sasprogressBar.value = 0
            if not os.path.exists(self.sml_data['db_file']):
                empty_db = os.path.join(os.path.dirname(__file__),'empty.sqlite')
                if os.path.exists(empty_db):
                    shutil.copy2(empty_db,self.sml_data['db_file'])
                else:
                    UniumPlugin.create_db(self.sml_data['db_file'])
                QgsMessageLog.logMessage(u'Новая база создана', level=QgsMessageLog.INFO)
            self.src_info["database"] = self.sml_data['db_file']
            self.dockwidget.sasprogressBar.setValue(25)
            qApp.processEvents()

            self.sml_data['table'] = UniumPlugin.create_table(self.sml_data['db_file'],self.sml_data['table'])
            self.src_info["datatable"] = self.sml_data['table']
            QgsMessageLog.logMessage(u'Новая таблица в базе создана', level=QgsMessageLog.INFO)
            self.dockwidget.sasprogressBar.setValue(50)
            qApp.processEvents()

            QgsMessageLog.logMessage(u'Начинаю загружать метки', level=QgsMessageLog.INFO)

            marksf = open(self.sml_data['marks_file'],'r')
            tree = etree.XML(marksf.read().decode('cp1251'))
            marksf.close()

            nodes = tree.xpath('/DATAPACKET/ROWDATA/ROW')

            uri = QgsDataSourceURI()
            uri.setDatabase(self.sml_data['db_file'])
            uri.setDataSource('', self.sml_data['table'], 'shape')

            lyr = QgsVectorLayer(uri.uri(),self.sml_data['table'],'spatialite')
            pr = lyr.dataProvider()
            lyr.startEditing()

            nodes_count = len(nodes)
            groups = [self.config["write_block"] for i in xrange(nodes_count/self.config["write_block"])]
            groups.append(nodes_count%self.config["write_block"])

            prgrs_interval = int(math.ceil(nodes_count/25.0))

            features = []
            group_idx = 0
            f_idx = 0
            position = 50

            for node in nodes:
                feature = QgsFeature()
                feature.setGeometry(QgsGeometry.fromPoint(QgsPoint(float(node.get('lonL')),float(node.get('LatB')))))
                fields = pr.fields()
                feature.setFields(fields)
                feature.setAttribute('name',node.get('name'))
                feature.setAttribute('descr',node.get('descr'))
                feature.setAttribute('sign',self.get_sign_by_src(node.get('picname')))
                feature.setAttribute('cat_id',int(node.get('categoryid')))
                features.append(feature)
                if len(features) == groups[group_idx]:
                    pr.addFeatures(features)
                    features = []
                    group_idx+=1
                f_idx+=1
                if f_idx%prgrs_interval == 0:
                    position+=1
                    self.dockwidget.sasprogressBar.setValue(position)
                    qApp.processEvents()

            lyr.commitChanges()
            self.dockwidget.sasprogressBar.setValue(75)
            qApp.processEvents()
            
            lyr = None
            
            QgsMessageLog.logMessage(u'Метки записаны в БД', level=QgsMessageLog.INFO)
            QgsMessageLog.logMessage(u'Начинаем добавлять слои', level=QgsMessageLog.INFO)
        
            catf = open(self.sml_data['cat_file'],'r')
            tree = etree.XML(catf.read().decode('cp1251'))
            catf.close()
            
            nodes = tree.xpath('/DATAPACKET/ROWDATA/ROW')
            
            root = QgsProject.instance().layerTreeRoot()
            root.removeAllChildren()
            QgsMapLayerRegistry.instance().removeAllMapLayers()

            self.categories = {}
            
            prgrs_interval = int(math.ceil(len(nodes)/25.0))

            f_idx = 0
            position = 75

            for node in nodes:
                chain = node.get('name').split(chr(92))
                self.categories[node.get('id')] = node.get('name')

                # Create sublayers for category
                c_root = UniumPlugin.create_sublayers(root, chain)

                # Create category vector layer
                cat_lyr = self.create_catlyr(uri.uri(),chain[len(chain)-1],int(node.get('id')))
                self.layers[cat_lyr.id()] = {'name': cat_lyr.name(),
                                    'subset': cat_lyr.subsetString(),
                                    'path': chr(92).join(chain[:len(chain)-1]),
                                    'full_name':chr(92).join(chain)}
                QgsMapLayerRegistry.instance().addMapLayer(cat_lyr,False)                            
                c_root.addLayer(cat_lyr)
                self.iface.legendInterface().setLayerVisible(cat_lyr, False)
                f_idx+=1
                if f_idx%prgrs_interval == 0:
                    position+=1
                    self.dockwidget.sasprogressBar.setValue(position)
                    qApp.processEvents()
                
            #self.get_layers()
            self.remove_vlayers_as_group()
            self.set_project_settings()
            canvas = self.iface.mapCanvas()
            renderer = canvas.mapRenderer()
            renderer.setDestinationCrs(self.mercator)
            canvas.setMapUnits(0)
            canvas.refresh()
            self.set_style_to_lyrs()
            renderer.setDestinationCrs(self.mercator)
            canvas.refresh()
            QgsMessageLog.logMessage(u'Слои созданы', level=QgsMessageLog.INFO)
            self.dockwidget.sasprogressBar.setValue(100)
            qApp.processEvents()
        except Exception, err:
            self.iface.messageBar().pushMessage("Error", u"Что-то случилось: %s" % err, level=QgsMessageBar.CRITICAL, duration=7)
            QgsMessageLog.logMessage(u'%s' % err, level=QgsMessageLog.CRITICAL)
    
    ##### Events methods #####
    def select_cat_file(self):
        self.sml_data['cat_file'] = QFileDialog.getOpenFileName(self.dockwidget, "Select Categorymarks.sml file ", os.path.dirname(self.sml_data.get('marks_file','')),"*.sml")
        self.dockwidget.catfileEdit.setText(self.sml_data['cat_file'])
    
    def select_marks_file(self):
        self.sml_data['marks_file'] = QFileDialog.getOpenFileName(self.dockwidget, "Select marks.sml file ", os.path.dirname(self.sml_data.get('cat_file','')),"*.sml")
        self.dockwidget.marksfileEdit.setText(self.sml_data['marks_file'])
        
    def select_db_file(self):
        self.sml_data['db_file'] = QFileDialog.getSaveFileName(self.dockwidget, "Select SpatiaLite database file ", os.path.dirname(self.sml_data.get('cat_file','')),"*.sqlite",QFileDialog.DontConfirmOverwrite)
        self.dockwidget.dbfileEdit.setText(self.sml_data['db_file'])
        
    def loadSASButton_clicked(self):
        self.sml_data['table'] = self.dockwidget.tbldbEdit.text()
        if self.sml_data.get('cat_file',None) and self.sml_data.get('marks_file',None) and self.sml_data.get('db_file', None) and self.sml_data.get('table', None):
            self.ParseSML()
            self.update_layers_list()

    def select_filefolder_clicked(self):
        sel_folder = QFileDialog.getExistingDirectory(self.dockwidget, "Select folder for saving files", self.config.get('files_folder',os.getenv("HOME")),QFileDialog.ShowDirsOnly)
        self.dockwidget.filefolderEdit.setText(sel_folder)
        self.config['files_folder'] = sel_folder

    @pyqtSlot(int)
    def wrblkBox_value_changed(self):
        self.config['write_block'] = self.dockwidget.wrblkBox.value()

    @pyqtSlot(int)
    def selected_layer_changed(self, idx):
        for lid in self.layers.keys():
            if self.layers[lid].get('full_name','') == self.dockwidget.layersBox.currentText():
                self.selected_id = lid
                self.dockwidget.nameEdit.setText(self.layers[lid].get('name_filter',''))
                self.dockwidget.descrEdit.setPlainText(self.layers[lid].get('descr_filter',''))
                QgsMessageLog.logMessage(u'Текущий слой: %s' % lid, level=QgsMessageLog.INFO)
                self.update_TView()

    def loadsetButton_clicked(self):
        self.getSettings()
        self.get_project_settings()
        self.updateSettingsUI()

    def savesetButton_clicked(self):
        self.set_project_settings()
        self.setSettings()

    def applyFilterButton_clicked(self):
        name_filter = self.dockwidget.nameEdit.text()
        descr_filter = self.dockwidget.descrEdit.toPlainText()
        lyrs = self.iface.legendInterface().layers()
        for lyr in lyrs:
            all_case = (self.selected_id == '0' and self.iface.legendInterface().isLayerVisible(lyr))
            if isinstance(lyr, QgsVectorLayer) and (lyr.id() == self.selected_id or all_case):
                self.set_subsets(lyr,name_filter,descr_filter)
                extent = lyr.extent()
                transf = QgsCoordinateTransform(lyr.crs(), self.iface.mapCanvas().mapSettings().destinationCrs())
                self.iface.mapCanvas().setExtent(transf.transform(extent))
                #self.iface.mapCanvas().zoomWithCenter(center.x(),center.y(), True)
                self.iface.mapCanvas().refresh()
                self.update_TView()

    def resetFilterButton_clicked(self):
        self.dockwidget.nameEdit.setText('')
        self.dockwidget.descrEdit.setPlainText('')
        lyrs = self.iface.legendInterface().layers()
        for lyr in lyrs:
            all_case = (self.selected_id == '0' and self.iface.legendInterface().isLayerVisible(lyr))
            if isinstance(lyr, QgsVectorLayer) and (lyr.id() == self.selected_id or all_case):
                self.reset_subsets(lyr)
                extent = lyr.extent()
                transf = QgsCoordinateTransform(lyr.crs(), self.iface.mapCanvas().mapSettings().destinationCrs())
                self.iface.mapCanvas().setExtent(transf.transform(extent))
                self.iface.mapCanvas().refresh()
                self.update_TView()

    def select_out_excel_file(self):
        xls_file = QFileDialog.getSaveFileName(self.dockwidget, "Select Excel file ", '',"Excel files (*.xlsx)")
        self.dockwidget.xlsoutEdit.setText(xls_file)

    def select_in_excel_file(self):
        xls_file = QFileDialog.getOpenFileName(self.dockwidget, "Select Excel file ", '',"Excel files (*.xlsx)")
        self.dockwidget.xlsinEdit.setText(xls_file)